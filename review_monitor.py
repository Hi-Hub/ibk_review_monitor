# -*- coding: utf-8 -*-
"""
IBK i-ONE Bank 5개 앱 - 부정 리뷰 모니터링 프로그램
=====================================================
매일 실행하면 아래 5개 앱(iOS + Android, 총 10개 채널)의 "부정적인 리뷰만"
모아서 최신순으로 정리한 엑셀 파일을 만들어줍니다.
맨 앞 탭에는 앱별·날짜별 신규 건수를 한눈에 볼 수 있는 요약표가 포함됩니다.

대상 앱: 개인고객용 / 기업고객용 / 알림 / 미니 / Global

사용법은 README.md 파일을 참고하세요.
"""

import json
import re
import time
import traceback
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from google_play_scraper import reviews as gp_reviews, Sort as GPSort
    GOOGLE_PLAY_AVAILABLE = True
except ImportError:
    GOOGLE_PLAY_AVAILABLE = False

# 엑셀(XML)이 허용하지 않는 제어문자/깨진 서로게이트 문자를 제거하기 위한 패턴
# (실제 앱스토어 리뷰에는 이모지 깨짐, 보이지 않는 제어문자 등이 섞여있는 경우가 있음)
ILLEGAL_CHARACTERS_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff\ufdd0-\ufdef\ufffe\uffff]"
)


def sanitize_text(value):
    """엑셀에 저장할 수 없는 문자를 제거합니다. (IllegalCharacterError 방지)"""
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)


# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
STATE_PATH = BASE_DIR / "data" / "state.json"
HISTORY_CSV = BASE_DIR / "data" / "전체_부정리뷰_누적기록.csv"
REPORTS_DIR = BASE_DIR / "reports"
LOGS_DIR = BASE_DIR / "logs"
LOG_FILE = LOGS_DIR / "실행로그.txt"
DOCS_DIR = BASE_DIR / "docs"
APP_DATA_PATH = DOCS_DIR / "data.json"

# 탭 순서: 사용자가 요청한 순서를 그대로 따름 (개인 -> 기업 -> 알림 -> 미니 -> 글로벌)
SHEET_ORDER = ["개인", "기업", "알림", "미니", "글로벌"]
PLATFORM_PREFIX = {"Android": "안드로이드", "iOS": "IOS"}
SUMMARY_SHEET_NAME = "일자별건수요약"


# ---------------------------------------------------------------------------
def log(message):
    """콘솔과 로그 파일에 동시에 메시지를 남깁니다."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_state():
    """이전에 확인한 리뷰 ID 목록을 불러옵니다.
    파일이 없거나 손상된 경우, 빈 상태로 새로 시작합니다."""
    if STATE_PATH.exists():
        try:
            with open(STATE_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                if "seen_ids" not in data:
                    data["seen_ids"] = {}
                return data
        except Exception as e:
            log(f"[경고] state.json 파일이 손상되어 새로 시작합니다: {e}")
    return {"seen_ids": {}}


def save_state(state):
    """state.json을 안전하게 저장합니다 (임시파일에 먼저 쓴 뒤 교체하여,
    저장 도중 프로그램이 종료되어도 파일이 깨지지 않도록 합니다)."""
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = STATE_PATH.with_suffix(".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    tmp_path.replace(STATE_PATH)


# ---------------------------------------------------------------------------
def fetch_ios_reviews(app_id, country, max_pages=10):
    """App Store 공식 리뷰 RSS(JSON)에서 최신 리뷰를 가져옵니다. (별도 인증 불필요)"""
    results = []
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    for page in range(1, max_pages + 1):
        url = (
            f"https://itunes.apple.com/{country}/rss/customerreviews/"
            f"page={page}/id={app_id}/sortby=mostrecent/json"
        )
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            log(f"  - iOS {page}페이지 요청 실패(건너뜀): {e}")
            time.sleep(2)
            continue

        entries = data.get("feed", {}).get("entry", [])
        if not entries:
            break  # 더 이상 리뷰 페이지가 없음

        page_had_review = False
        for item in entries:
            if "author" not in item:  # 앱 정보 항목(리뷰 아님)은 건너뜀
                continue
            page_had_review = True
            try:
                results.append({
                    "platform": "iOS",
                    "review_id": item["id"]["label"],
                    "date": item["updated"]["label"][:19].replace("T", " "),
                    "rating": int(item["im:rating"]["label"]),
                    "author": item["author"]["name"]["label"],
                    "title": item.get("title", {}).get("label", ""),
                    "content": item.get("content", {}).get("label", ""),
                })
            except Exception as e:
                log(f"  - iOS 리뷰 항목 파싱 오류(건너뜀): {e}")

        if not page_had_review:
            break
        time.sleep(0.5)

    return results


def fetch_android_reviews(package_name, lang="ko", country="kr", count=200):
    """Google Play 공개 리뷰를 가져옵니다. (별도 인증 불필요)"""
    if not GOOGLE_PLAY_AVAILABLE:
        log("  - google-play-scraper 패키지가 설치되어 있지 않아 Android 리뷰는 건너뜁니다.")
        log("    (setup.bat 을 먼저 실행해주세요)")
        return []

    results = []
    try:
        items, _ = gp_reviews(
            package_name, lang=lang, country=country,
            sort=GPSort.NEWEST, count=count,
        )
    except Exception as e:
        log(f"  - Android 리뷰 요청 실패: {e}")
        return []

    for item in items:
        try:
            at = item.get("at")
            date_str = at.strftime("%Y-%m-%d %H:%M:%S") if at else ""
            results.append({
                "platform": "Android",
                "review_id": str(item["reviewId"]),
                "date": date_str,
                "rating": int(item["score"]),
                "author": item.get("userName", "") or "",
                "title": "",
                "content": item.get("content", "") or "",
            })
        except Exception as e:
            log(f"  - Android 리뷰 항목 파싱 오류(건너뜀): {e}")

    return results


# ---------------------------------------------------------------------------
def classify_complaint(text, categories, default_category):
    """리뷰 내용에 포함된 키워드를 기준으로 불만사항 카테고리를 분류합니다.
    (참고: 키워드 매칭 방식이며, AI 의미분석이 아니므로 100% 정확하지는 않습니다.
    config.json 의 complaint_categories 항목에서 키워드를 자유롭게 보완할 수 있습니다.)"""
    if not text:
        return default_category
    for cat in categories:
        for kw in cat["keywords"]:
            if kw in text:
                return cat["name"]
    return default_category


def format_date_kr(date_str):
    """'YYYY-MM-DD HH:MM:SS' 형식의 날짜를 'MM월 DD일' 형식으로 변환합니다."""
    try:
        dt = datetime.strptime(date_str[:19], "%Y-%m-%d %H:%M:%S")
        return f"{dt.month:02d}월 {dt.day:02d}일"
    except Exception:
        return date_str


def append_history(rows):
    """이번 실행에서 새로 확인된(=신규) 부정 리뷰를 누적 CSV 파일에 추가 기록합니다."""
    if not rows:
        return
    HISTORY_CSV.parent.mkdir(parents=True, exist_ok=True)
    is_new_file = not HISTORY_CSV.exists()
    with open(HISTORY_CSV, "a", encoding="utf-8-sig", newline="") as f:
        if is_new_file:
            f.write("수집일시,앱,플랫폼,작성일,평점,사용자ID,리뷰,불만사항\n")
        collected_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for r in rows:
            content = r["content"].replace('"', "'").replace("\n", " ")
            author = r["author"].replace('"', "'")
            f.write(
                f'{collected_at},{r["app_key"]},{r["platform"]},{r["date_sort"]},'
                f'{r["rating"]},"{author}","{content}",{r["category"]}\n'
            )


# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="2E5C8A")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
TAB_TITLE_FONT = Font(name="맑은 고딕", bold=True, size=16, color="1F3B66")
SUB_TITLE_FONT = Font(name="맑은 고딕", size=11, color="44546A")
SUMMARY_FONT = Font(name="맑은 고딕", size=10, color="555555")
ERROR_FONT = Font(name="맑은 고딕", size=10, color="B23B3B", italic=True)
TOTAL_FILL = PatternFill("solid", start_color="EEF3FA")
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = ["일시", "사용자ID", "리뷰", "불만사항"]
COL_WIDTHS = {"A": 12, "B": 18, "C": 70, "D": 16}


def build_sheet(wb, sheet_name, app_label, platform, rows, error_note=None):
    """앱 1개 x 플랫폼 1개에 해당하는 시트를 만듭니다. (부정 리뷰만, 최신순 정렬)"""
    ws = wb.create_sheet()
    ws.title = sheet_name

    # 1행: 탭 이름 자체를 큰 제목으로 표시
    ws["A1"] = sanitize_text(sheet_name)
    ws["A1"].font = TAB_TITLE_FONT

    # 2행: 앱 설명(어떤 앱/플랫폼인지)
    ws["A2"] = sanitize_text(f"{app_label} ({platform})")
    ws["A2"].font = SUB_TITLE_FONT

    # 3행: 요약 통계
    new_count = sum(1 for r in rows if r.get("is_new"))
    if rows:
        summary = f"부정 리뷰 {len(rows)}건  |  그 중 신규 {new_count}건"
    else:
        summary = "수집된 부정 리뷰 없음"
    ws["A3"] = summary
    ws["A3"].font = SUMMARY_FONT

    if error_note:
        ws["A4"] = f"[오류] {error_note}"
        ws["A4"].font = ERROR_FONT
        header_row = 6
    else:
        header_row = 5

    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 최신 리뷰가 맨 위로 오도록 정렬
    sorted_rows = sorted(rows, key=lambda r: r["date_sort"], reverse=True)

    for i, r in enumerate(sorted_rows):
        excel_row = header_row + 1 + i
        values = [r["date_disp"], r["author"], r["content"], r["category"]]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=sanitize_text(v))
            cell.border = BORDER
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if col_idx in (1, 4) else "left",
                wrap_text=(col_idx == 3),
            )

    last_row = max(header_row + 1, header_row + len(sorted_rows))
    ws.auto_filter.ref = f"A{header_row}:D{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _col_letter(idx):
    """1=A, 2=B, ... 26=Z, 27=AA 식으로 엑셀 컬럼 문자를 만듭니다."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def build_daily_count_summary_sheet(wb, channel_names, counts_by_date):
    """앱별·날짜별 '부정 리뷰' 건수를 한 장으로 정리한 요약 탭을 만듭니다.
    (그날 작성된 부정 리뷰의 전체 건수이며, 신규 여부와는 무관합니다.)

    counts_by_date: { 'YYYY-MM-DD': { 채널명: 건수 } } 형태의 데이터
    channel_names: 표에 표시할 채널(탭) 이름 목록, 표시 순서 그대로
    """
    ws = wb.active
    ws.title = SUMMARY_SHEET_NAME

    ws["A1"] = SUMMARY_SHEET_NAME
    ws["A1"].font = TAB_TITLE_FONT
    ws["A2"] = "작성일자 기준 · 앱별 부정 리뷰 건수 (이번 실행에서 수집된 전체 부정 리뷰 기준)"
    ws["A2"].font = SUB_TITLE_FONT

    header_row = 4
    headers = ["날짜"] + channel_names + ["합계"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sorted_dates = sorted(counts_by_date.keys(), reverse=True)  # 최신 날짜가 위로

    col_totals = defaultdict(int)
    row_idx = header_row + 1
    for date_key in sorted_dates:
        row_counts = counts_by_date[date_key]
        try:
            date_disp = datetime.strptime(date_key, "%Y-%m-%d").strftime("%m월 %d일")
        except Exception:
            date_disp = date_key

        row_total = 0
        date_cell = ws.cell(row=row_idx, column=1, value=date_disp)
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.border = BORDER
        for col_idx, ch in enumerate(channel_names, start=2):
            count = row_counts.get(ch, 0)
            row_total += count
            col_totals[ch] += count
            cell = ws.cell(row=row_idx, column=col_idx, value=count if count else "")
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
        total_cell = ws.cell(row=row_idx, column=len(headers), value=row_total if row_total else "")
        total_cell.alignment = Alignment(horizontal="center")
        total_cell.font = Font(bold=True)
        total_cell.border = BORDER
        row_idx += 1

    if not sorted_dates:
        ws.cell(row=row_idx, column=1, value="수집된 부정 리뷰가 없습니다.")
        row_idx += 1

    # 맨 아래 합계 행
    grand_total = sum(col_totals.values())
    total_row_label = ws.cell(row=row_idx, column=1, value="합계")
    total_row_label.font = Font(bold=True)
    total_row_label.fill = TOTAL_FILL
    total_row_label.border = BORDER
    total_row_label.alignment = Alignment(horizontal="center")
    for col_idx, ch in enumerate(channel_names, start=2):
        cell = ws.cell(row=row_idx, column=col_idx, value=col_totals[ch] if col_totals[ch] else "")
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    grand_cell = ws.cell(row=row_idx, column=len(headers), value=grand_total if grand_total else "")
    grand_cell.font = Font(bold=True)
    grand_cell.fill = TOTAL_FILL
    grand_cell.border = BORDER
    grand_cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"B{header_row + 1}"
    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[_col_letter(col_idx)].width = 14


# ---------------------------------------------------------------------------
def process_channel(app_key, app_cfg, platform, config, seen_ids_before):
    """채널(앱 x 플랫폼) 1개를 수집·가공합니다. 부정 리뷰만 결과에 포함됩니다.
    반환값: (raw 전체 리스트, 부정 리뷰 행 리스트, 이번에 확인한 전체 review_id 집합)"""
    categories = config.get("complaint_categories", [])
    default_category = config.get("default_category", "기타")
    neg_threshold = config.get("negative_max_rating", 2)

    if platform == "iOS":
        raw = fetch_ios_reviews(
            app_cfg["ios_app_id"], config.get("ios_country", "kr"),
            max_pages=config.get("ios_max_pages", 10),
        )
    else:
        raw = fetch_android_reviews(
            app_cfg["android_package"],
            lang=config.get("android_lang", "ko"),
            country=config.get("android_country", "kr"),
            count=config.get("android_max_count", 200),
        )

    rows = []
    current_ids = set()

    for r in raw:
        rid = r["review_id"]
        current_ids.add(rid)

        if r["rating"] > neg_threshold:
            continue  # 긍정/중립 리뷰는 저장하지 않음

        is_new = rid not in seen_ids_before

        content_full = (r["content"] or "").strip()
        if r.get("title"):
            content_full = f"[{r['title']}] {content_full}".strip()
        content_full = sanitize_text(content_full)
        author_clean = sanitize_text(r["author"] or "(익명)")

        category = classify_complaint(content_full, categories, default_category)

        rows.append({
            "app_key": app_key,
            "platform": platform,
            "date_disp": format_date_kr(r["date"]),
            "date_sort": r["date"],
            "rating": r["rating"],
            "author": author_clean,
            "content": content_full,
            "category": category,
            "is_new": is_new,
        })

    return raw, rows, current_ids


# ---------------------------------------------------------------------------
def export_app_data(channel_sheet_data, counts_by_date, channel_names_in_order):
    """폰 앱(PWA)이 읽어들일 data.json 파일을 만듭니다.
    엑셀과 내용은 같지만, 웹 화면에서 그대로 그리기 쉬운 JSON 구조로 저장합니다."""
    sorted_dates = sorted(counts_by_date.keys(), reverse=True)
    date_rows = []
    channel_totals = defaultdict(int)
    grand_total = 0
    for date_key in sorted_dates:
        row_counts = counts_by_date[date_key]
        try:
            date_disp = datetime.strptime(date_key, "%Y-%m-%d").strftime("%m월 %d일")
        except Exception:
            date_disp = date_key
        row_total = 0
        counts = {}
        for ch in channel_names_in_order:
            c = row_counts.get(ch, 0)
            counts[ch] = c
            row_total += c
            channel_totals[ch] += c
        grand_total += row_total
        date_rows.append({"date": date_disp, "counts": counts, "total": row_total})

    channels_out = []
    for sheet_name, app_label, platform, rows, error_note in channel_sheet_data:
        sorted_rows = sorted(rows, key=lambda r: r["date_sort"], reverse=True)
        new_count = sum(1 for r in rows if r.get("is_new"))
        channels_out.append({
            "key": sheet_name,
            "app_label": app_label,
            "platform": platform,
            "error": error_note,
            "total": len(rows),
            "new_count": new_count,
            "reviews": [
                {
                    "date_disp": r["date_disp"],
                    "date_sort": r["date_sort"],
                    "rating": r["rating"],
                    "author": r["author"],
                    "content": r["content"],
                    "category": r["category"],
                    "is_new": r["is_new"],
                }
                for r in sorted_rows
            ],
        })

    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {
            "channels": channel_names_in_order,
            "dates": date_rows,
            "channel_totals": channel_totals,
            "grand_total": grand_total,
        },
        "channels": channels_out,
    }

    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    tmp_path = APP_DATA_PATH.with_suffix(".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    tmp_path.replace(APP_DATA_PATH)
    log(f"폰 앱용 데이터 생성 완료: {APP_DATA_PATH}")


# ---------------------------------------------------------------------------
def main():
    log("=" * 60)
    log("IBK i-ONE Bank 부정 리뷰 모니터링 시작")

    config = load_config()
    state = load_state()
    seen = state.get("seen_ids", {})

    apps_by_key = {a["key"]: a for a in config["apps"]}

    channel_sheet_data = []  # [(sheet_name, app_label, platform, rows, error_note), ...]
    counts_by_date = defaultdict(lambda: defaultdict(int))  # {날짜: {채널명: 전체건수}}
    channel_names_in_order = []

    history_rows = []
    grand_total_negative = 0
    grand_total_new = 0

    for app_key in SHEET_ORDER:
        app_cfg = apps_by_key.get(app_key)
        if not app_cfg:
            log(f"[경고] config.json 에 '{app_key}' 앱 설정이 없어 건너뜁니다.")
            continue

        for platform in ["Android", "iOS"]:
            sheet_name = f"{PLATFORM_PREFIX[platform]}{app_key}"
            channel_names_in_order.append(sheet_name)
            log(f"[{sheet_name}] 리뷰 수집 중...")

            try:
                state_key = f"{app_key}_{platform}"
                seen_ids_before = set(seen.get(state_key, []))

                raw, rows, current_ids = process_channel(
                    app_key, app_cfg, platform, config, seen_ids_before
                )
                log(f"  -> 전체 {len(raw)}건 조회됨")

                # 이번에 수집된 모든 리뷰 ID를 "확인 완료"로 기록합니다.
                seen[state_key] = list(seen_ids_before | current_ids)
                state["seen_ids"] = seen

                # 채널 하나를 처리할 때마다 즉시 저장합니다.
                # (뒤에 나오는 다른 채널에서 오류가 발생해도,
                #  이미 처리된 채널의 "확인 완료" 기록은 절대 사라지지 않습니다.)
                save_state(state)

                new_rows = [r for r in rows if r["is_new"]]
                grand_total_negative += len(rows)
                grand_total_new += len(new_rows)
                log(f"  -> 부정 리뷰 {len(rows)}건 (그 중 신규 {len(new_rows)}건)")

                for r in rows:
                    date_key = r["date_sort"][:10]
                    counts_by_date[date_key][sheet_name] += 1

                history_rows.extend(new_rows)
                channel_sheet_data.append((sheet_name, app_cfg["label"], platform, rows, None))

            except Exception:
                # 한 채널에서 오류가 나더라도 전체 프로그램이 멈추지 않고,
                # 나머지 채널은 계속 정상적으로 진행되도록 합니다.
                log(f"  - [오류] [{sheet_name}] 처리 중 문제가 발생하여 이 채널은 건너뜁니다.")
                with open(LOG_FILE, "a", encoding="utf-8") as f:
                    f.write(f"[{datetime.now()}] [{sheet_name}] 오류 상세:\n{traceback.format_exc()}\n")
                channel_sheet_data.append((
                    sheet_name, app_cfg["label"], platform, [],
                    "이 채널 처리 중 오류가 발생했습니다. logs/실행로그.txt 를 확인해주세요.",
                ))
                continue

    # ---- 엑셀 조립: 맨 앞에 요약 탭, 그 다음 채널별 탭 ----
    wb = Workbook()
    build_daily_count_summary_sheet(wb, channel_names_in_order, counts_by_date)

    for sheet_name, app_label, platform, rows, error_note in channel_sheet_data:
        build_sheet(wb, sheet_name, app_label, platform, rows, error_note=error_note)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    today_str = datetime.now().strftime("%Y%m%d")
    output_path = REPORTS_DIR / f"부정리뷰리포트_{today_str}.xlsx"
    wb.save(output_path)
    log(f"엑셀 리포트 생성 완료: {output_path}")

    export_app_data(channel_sheet_data, counts_by_date, channel_names_in_order)

    append_history(history_rows)

    log(f"전체 부정 리뷰: {grand_total_negative}건 / 신규 부정 리뷰: {grand_total_new}건")
    log("작업 완료")
    log("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now()}] 오류 발생:\n{traceback.format_exc()}\n")
        print("\n[오류] 프로그램 실행 중 문제가 발생했습니다.")
        print("logs/실행로그.txt 파일에서 자세한 내용을 확인하거나,")
        print("인터넷 연결 상태를 확인한 뒤 다시 실행해주세요.\n")
